"""xlsx → DB 적재 어댑터 (레벨 일반화).

사용법:
    poetry run python scripts/load_chunks.py --level N5

`data/{level}/{level}_master.xlsx` (5시트) + `{level}_comparison.xlsx` (2시트)를 읽어
`chunks` / `comparison_pairs` 테이블에 UPSERT 한다. N4·N3도 동일 명령으로 처리.

설계 계약: `docs/planning/session_4/be_jaehyeon/04_target_db_design.md` §3.
- 어댑터가 유일한 정규화 지점 (source_pool 확장 / 미결→None / star 정수화 등).
- comparison_pair_ids는 저장하지 않음 (comparison_pairs에서 도출, 단방향).
- embedding 벡터는 별도 단계(embed_chunks.py)에서 채움 — 여기선 텍스트까지만.
"""

from __future__ import annotations

import argparse
import asyncio
import os
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

# 프로젝트 루트를 import 경로에 추가 (직접 실행 대비)
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.core.config import settings  # noqa: E402
from src.db.models import Chunk, ComparisonPair  # noqa: E402


# ── 정규화 함수 (어댑터 계약 §3-2) ──

def expand_source_pool(raw: str | None) -> list[str]:
    """`"SRC-01,02,03"` → `["SRC-01","SRC-02","SRC-03"]` (첫 토큰 외 SRC- 재접두)."""
    if not raw:
        return []
    out: list[str] = []
    for tok in str(raw).split(","):
        t = tok.strip()
        if not t:
            continue
        out.append(t if t.startswith("SRC-") else f"SRC-{t}")
    return out


def to_bool_or_none(value: Any) -> bool | None:
    """`"True"/"False"/"미결"` → `True/False/None` (미결·미평가 = None)."""
    if value is None:
        return None
    s = str(value).strip().lower()
    if s in ("true", "1"):
        return True
    if s in ("false", "0"):
        return False
    return None


def star_to_int(value: Any) -> int:
    """`"3star"` → `3`."""
    s = str(value).strip().lower().replace("star", "").strip()
    return int(s)


# ── xlsx 읽기 ──

def _sheet_rows(ws: Any) -> list[dict[str, Any]]:
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return []
    header = list(data[0])
    return [
        dict(zip(header, row))
        for row in data[1:]
        if any(cell is not None for cell in row)
    ]


def build_payloads(
    level: str, master_path: Path, comparison_path: Path
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """xlsx → (chunks payload, comparison_pairs payload)."""
    wb = load_workbook(master_path, read_only=True)
    master = {r["grammar_point_id"]: r for r in _sheet_rows(wb["master_list"])}
    l3 = {r["grammar_point_id"]: r for r in _sheet_rows(wb["l3_assignment"])}
    border = {r["grammar_point_id"]: r for r in _sheet_rows(wb["border_meta"])}
    variants = _sheet_rows(wb["variant_chunks"])
    wb.close()

    wb2 = load_workbook(comparison_path, read_only=True)
    pairs = _sheet_rows(wb2["comparison_pairs"])
    comp_chunks = {r["compare_id"]: r for r in _sheet_rows(wb2["comparison_chunks"])}
    wb2.close()

    chunks: list[dict[str, Any]] = []

    # point
    for gid, m in master.items():
        tags = l3.get(gid, {})
        chunks.append(
            {
                "chunk_key": gid,
                "level": level,
                "chunk_type": "point",
                "grammar_point_id": gid,
                "base_point_id": None,
                "border_flag": to_bool_or_none(tags.get("border_candidate")),
                "body": {
                    "japanese_name": m.get("japanese_name"),
                    "korean_meaning": m.get("korean_meaning"),
                    "l1": m.get("l1"),
                    "l2": m.get("l2"),
                    "source_pool": expand_source_pool(m.get("source_pool")),
                    "frequency": m.get("frequency"),
                    "note": m.get("note"),
                },
                "l3_tags": {
                    "connection_type": tags.get("connection_type"),
                    "tense": tags.get("tense"),
                    "polarity": tags.get("polarity"),
                    "formality": tags.get("formality"),
                }
                if tags
                else None,
                "border_meta": {
                    "border_reason": border[gid].get("border_reason"),
                    "n5_scope": border[gid].get("n5_scope"),
                    "n4_advanced": border[gid].get("n4_advanced"),
                }
                if gid in border
                else None,
                "embedding_text": tags.get("embedding_text") or "",
                "source_status": "validated",
            }
        )

    # variant
    for v in variants:
        chunks.append(
            {
                "chunk_key": v["grammar_point_id"],
                "level": v.get("level") or level,
                "chunk_type": "variant",
                "grammar_point_id": v["grammar_point_id"],
                "base_point_id": v.get("base_id"),
                "border_flag": to_bool_or_none(v.get("border_candidate")),
                "body": {
                    "japanese_name": v.get("japanese_name"),
                    "korean_meaning": v.get("korean_meaning"),
                    "variant_label": v.get("variant_label"),
                    "l1": v.get("l1"),
                    "l2": v.get("l2"),
                    "note": v.get("note"),
                },
                "l3_tags": None,
                "border_meta": None,
                "embedding_text": v.get("embedding_text") or "",
                "source_status": "validated",
            }
        )

    # compare (comparison_chunks 본문 + comparison_pairs 혼동요점)
    pair_by_id = {p["compare_id"]: p for p in pairs}
    for cid, cc in comp_chunks.items():
        p = pair_by_id.get(cid, {})
        chunks.append(
            {
                "chunk_key": cid,
                "level": level,
                "chunk_type": "compare",
                "grammar_point_id": None,
                "base_point_id": None,
                "border_flag": None,
                "body": {
                    "pair_label": cc.get("pair_label"),
                    "confusion_point": p.get("confusion_point"),
                    "cluster": cc.get("cluster"),
                    "star_weight": cc.get("star_weight"),
                },
                "l3_tags": None,
                "border_meta": None,
                "embedding_text": cc.get("embedding_text") or "",
                "source_status": "validated",
            }
        )

    pair_payloads: list[dict[str, Any]] = [
        {
            "compare_id": p["compare_id"],
            "level": level,
            "left_point_id": p["left_grammar_point_id"],
            "right_point_id": p["right_grammar_point_id"],
            "cluster": p["cluster"],
            "star_weight": star_to_int(p["star_weight"]),
            "confusion_point": p.get("confusion_point") or "",
        }
        for p in pairs
    ]
    return chunks, pair_payloads


async def _upsert(
    chunks: list[dict[str, Any]], pairs: list[dict[str, Any]]
) -> None:
    engine = create_async_engine(settings.database_url)
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as session:
        for c in chunks:
            stmt = insert(Chunk).values(**c)
            stmt = stmt.on_conflict_do_update(
                index_elements=["chunk_key"],
                set_={k: stmt.excluded[k] for k in c if k != "chunk_key"},
            )
            await session.execute(stmt)
        for p in pairs:
            stmt = insert(ComparisonPair).values(**p)
            stmt = stmt.on_conflict_do_update(
                index_elements=["compare_id"],
                set_={k: stmt.excluded[k] for k in p if k != "compare_id"},
            )
            await session.execute(stmt)
        await session.commit()
    await engine.dispose()


def main() -> None:
    parser = argparse.ArgumentParser(description="xlsx → DB 적재 (레벨 일반화)")
    parser.add_argument("--level", required=True, help="N5 / N4 / N3 ...")
    args = parser.parse_args()

    level = args.level.upper()
    base = Path("data") / level.lower()
    master_path = base / f"{level.lower()}_master.xlsx"
    comparison_path = base / f"{level.lower()}_comparison.xlsx"
    for p in (master_path, comparison_path):
        if not p.exists():
            raise SystemExit(f"파일 없음: {p}")

    chunks, pairs = build_payloads(level, master_path, comparison_path)
    asyncio.run(_upsert(chunks, pairs))

    points = sum(1 for c in chunks if c["chunk_type"] == "point")
    variants = sum(1 for c in chunks if c["chunk_type"] == "variant")
    compares = sum(1 for c in chunks if c["chunk_type"] == "compare")
    print(
        f"[{level}] 적재 완료 — chunks={len(chunks)} "
        f"(point {points} / variant {variants} / compare {compares}), "
        f"comparison_pairs={len(pairs)}"
    )


if __name__ == "__main__":
    main()
